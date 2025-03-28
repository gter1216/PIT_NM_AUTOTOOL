#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   NIO PIT 自动化工具
   Author:  Charles Xu (charles.xu2@nio.com)
   Company: NIO
   Created: 2025-02-25 | Version: 1.0
-------------------------------------------------
   Copyright (c) 2025 Charles Xu (NIO)
   Licensed under the MIT License: 
   https://opensource.org/licenses/MIT
-------------------------------------------------
   Change Log:
   2025-02-25  Charles Xu  Initial creation
-------------------------------------------------
"""

import os
import sys
import re
import pandas as pd
import numpy as np
# 添加项目根目录到 Python 路径
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils import read_excel, save_excel
from parse_diag_table.config import get_column_index
from parse_diag_table.config import (
    SOURCE_SHEET1, SOURCE_SHEET2, SOURCE_SHEET3,
    TARGET_SHEET1, TARGET_SHEET2,
)

# 读取 Excel 文件并做预处理
def pre_process_31(source_file, target_file):
    # # 设置 pandas 显示选项，显示所有列
    # pd.set_option('display.max_columns', None)  # 显示所有列
    # pd.set_option('display.width', None)        # 加宽显示宽度
    # pd.set_option('display.max_colwidth', None) # 显示完整的列内容
    
    # print(f"Processing source file: {source_file}, sheet: {SOURCE_SHEET3}")
    source_df = read_excel(source_file, SOURCE_SHEET3)
    # print(f"\nSource DataFrame:")
    # print(f"总行数: {len(source_df)}")
    # print(source_df)
    
    # print(f"\nProcessing target file: {target_file}, sheet: {TARGET_SHEET2}")
    target_df = read_excel(target_file, TARGET_SHEET2)
    # print("\nTarget DataFrame:")
    # print(f"总行数: {len(target_df)}")
    # print(target_df)
    # print("\nTarget DataFrame Info:")
    # print(target_df.info())
    
    # 提取 source_rid_list
    source_rid_list = []
    for value in source_df.iloc[:, get_column_index('A', 'routine_control')].dropna().astype(str):
        if value.startswith("0x"):
            source_rid_list.append(value)
    print(f'source RID len: {len(source_rid_list)}; \nsource RID: {', '.join(source_rid_list)}')
    
    # 提取 target_rid_list，每三行为一组
    target_rid_list = {}
    current_rid = None
    current_description = None
    
    for idx, row in target_df.iterrows():
        rid_value = row.iloc[get_column_index('A', 'rid_library')]
        desc_value = row.iloc[get_column_index('B', 'rid_library')]
        
        # 检查是否为有效的 RID 值（不是空值且不是特殊字符串）
        if not pd.isna(rid_value) and str(rid_value).strip().upper() not in ['NAN', 'NA']:
            current_rid = rid_value
            current_description = desc_value
            target_rid_list[current_rid] = idx - (idx % 3)
        else:
            # 将前两列的值设置为与组内第一行相同
            target_df.iloc[idx, get_column_index('A', 'rid_library')] = current_rid
            target_df.iloc[idx, get_column_index('B', 'rid_library')] = current_description
    
    print(f'target RID len: {len(target_rid_list)}; \ntarget RID: {", ".join([f"{k}: {v}" for k, v in target_rid_list.items()])}')

    # for rid, start_idx in target_rid_list.items():
    #     print(f"RID: {rid}, 起始行号: {start_idx}")
    #     for i in range(3):
    #         row_data = target_df.iloc[start_idx + i]
    #         print(f"行 {start_idx + i} 数据: {row_data.to_dict()}")

    # 进行比对和更新
    new_rids = []
    removed_indices = []
    
    # 找出需要新增的 RID
    for value in source_rid_list:
        if value not in target_rid_list:
            print(f"新增: {value}")
            new_rids.append(value)
    
    # 找出需要删除的 RID
    for value, start_idx in sorted(target_rid_list.items(), key=lambda x: x[1], reverse=True):
        if value not in source_rid_list:
            print(f"删除: {value}")
            # 记录三行的索引
            removed_indices.extend([start_idx, start_idx + 1, start_idx + 2])
    

    # print("\n删除前的 DataFrame:")
    # print(f"总行数: {len(target_df)}")
    # print(target_df)

    # 删除不存在的 RID（每次删除三行）
    if removed_indices:
        # 按照索引从大到小排序，这样删除时不会影响其他行的索引
        removed_indices.sort(reverse=True)
        target_df.drop(removed_indices, inplace=True)
        target_df.reset_index(drop=True, inplace=True)
        
    # print("\n删除后的 DataFrame:")
    # print(f"总行数: {len(target_df)}")
    # print(target_df)
    
    # 添加新的 RID（每次添加三行）
    for rid in new_rids:
        # 创建三行新数据
        for i in range(3):
            new_row = ['NA'] * len(target_df.columns)  # 初始化所有列为 'NA'
            # 在所有三行中都设置相同的 RID 和 Description
            new_row[get_column_index('A', 'rid_library')] = rid
            target_df.loc[len(target_df)] = new_row

    # print("\n预处理后的 DataFrame:")
    # print(f"总行数: {len(target_df)}")
    # print(target_df)
    
    return source_df, target_df

def save_excel_with_merged_cells(df, file_path, sheet_name):
    """
    保存 DataFrame 到 Excel 文件，并调整列宽。
    如果 sheet 已存在，则覆盖该 sheet，保留其他 sheet。
    """
    try:
        # 第一步：使用 pandas 保存 DataFrame
        if os.path.exists(file_path):
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # 第二步：使用 openpyxl 打开文件进行单元格合并
        from openpyxl import load_workbook
        wb = load_workbook(file_path)
        ws = wb[sheet_name]
        
        # 调整列宽
        for idx, col in enumerate(df.columns):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            ws.column_dimensions[chr(65 + idx)].width = max_length + 2
        
        # 合并 A 列和 B 列的单元格
        current_rid = None
        merge_start = 2  # Excel 行号从 1 开始，第 1 行是标题
        
        for row in range(2, len(df) + 2):  # +2 是因为 Excel 行号从 1 开始，且第 1 行是标题
            rid = df.iloc[row-2, get_column_index('A', 'rid_library')]
            if current_rid != rid:
                if current_rid is not None:
                    # 合并前一个 RID 的三行
                    ws.merge_cells(f'A{merge_start}:A{row-1}')
                    ws.merge_cells(f'B{merge_start}:B{row-1}')
                current_rid = rid
                merge_start = row
        
        # 处理最后一个 RID
        if current_rid is not None:
            ws.merge_cells(f'A{merge_start}:A{len(df)+1}')
            ws.merge_cells(f'B{merge_start}:B{len(df)+1}')
        
        # 保存修改后的文件
        wb.save(file_path)
        
    except Exception as e:
        print(f"保存 Excel 文件时出错: {str(e)}")
        raise

def update_target_with_source_data(subservice,source_df, target_df, target_row_idx, matching_rows, rid_start_row):
    """当 source 中有匹配数据时，更新 target 数据
    Args:
        source_df: 源数据框
        target_df: 目标数据框
        target_row_idx: 目标行索引
        matching_rows: 源数据中匹配的行索引列表 (非空)
        rid_start_row: RID 在 source 中的起始行索引
    """
    # 更新 B 列 Description (使用 RID 起始行的 B 列值)
    target_df.iloc[target_row_idx, get_column_index('B', 'rid_library')] = \
        source_df.iloc[rid_start_row, get_column_index('B', 'routine_control')]
        
    # 更新 F 列 LockLevel (使用 RID 起始行的 D 列值)
    target_df.iloc[target_row_idx, get_column_index('F', 'rid_library')] = \
        source_df.iloc[rid_start_row, get_column_index('D', 'routine_control')]
    
    # 更新 G 列 Session (使用 RID 起始行的 E 列 Session 的值)
    target_df.iloc[target_row_idx, get_column_index('G', 'rid_library')] = \
        f"0x{source_df.iloc[rid_start_row, get_column_index('E', 'routine_control')]}"
    
    # 根据 source 的 E 列 Session 的值更新 C 列 APP 和 D 列 Boot
    session_value = str(source_df.iloc[rid_start_row, get_column_index('E', 'routine_control')])
    if session_value == '02':
        target_df.iloc[target_row_idx, get_column_index('C', 'rid_library')] = "N"
        target_df.iloc[target_row_idx, get_column_index('D', 'rid_library')] = "Y"
    elif session_value == '03':
        target_df.iloc[target_row_idx, get_column_index('C', 'rid_library')] = "Y"
        target_df.iloc[target_row_idx, get_column_index('D', 'rid_library')] = "N"
    else:
        raise ValueError(f"Session 值 {session_value} 无效")
    
    # 根据 当前 subservice 的值更新 E 列 SubService 的值，前面加个 0x
    target_df.iloc[target_row_idx, get_column_index('E', 'rid_library')] = f"0x{subservice}"

    # H 列的值 RequestData 直接更新为 NA
    target_df.iloc[target_row_idx, get_column_index('H', 'rid_library')] = "NA"

    # I 列的值 ResponseLength 比较复杂，需要根据 source H列的 Req/Resp 和 M 列的 BitLength 一起来考虑
    # 如果 H 列的值为 Resp, 则 M列的值相加处于8 就是 ResponseLength 的值
    # 如果 H 列的值为 Req，则忽略这行
    total_bits = 0
    resp_length_str = False
    for row in matching_rows:
        req_resp = str(source_df.iloc[row, get_column_index('H', 'routine_control')])
        if req_resp == 'Resp':
            bit_length = source_df.iloc[row, get_column_index('M', 'routine_control')]
            if pd.notna(bit_length):
                if isinstance(bit_length, str):
                    resp_length_str = True
                    break
                total_bits += bit_length
    # 计算 ResponseLength
    if resp_length_str:
        target_df.iloc[target_row_idx, get_column_index('I', 'rid_library')] = "NA"
    else:
        if total_bits > 0:
            target_df.iloc[target_row_idx, get_column_index('I', 'rid_library')] = int(total_bits / 8)
        else:
            target_df.iloc[target_row_idx, get_column_index('I', 'rid_library')] = "NA"
            raise ValueError(f"total_bits is 0, please check the source data")
    
    # J 列的值 ResponseData 直接更新为 NA
    target_df.iloc[target_row_idx, get_column_index('J', 'rid_library')] = "NA"

    # K 列的值 ResponseNRC 根据 source H列的 Req/Resp 和 M 列的 BitLength 一起来考虑
    # 如果 H 列的值为 Req，则 查看 M 列的 Bit length，只要有一个 Req 对应的 Bit length 不为 NA，则 K 列的值为 0x13，否则为 NA
    # 如果 H 列的值为 Resp，则 忽略
    target_df.iloc[target_row_idx, get_column_index('K', 'rid_library')] = "NA"
    for row in matching_rows:
        req_resp = str(source_df.iloc[row, get_column_index('H', 'routine_control')])
        if req_resp == 'Req':
            bit_length = source_df.iloc[row, get_column_index('M', 'routine_control')]
            if pd.notna(bit_length):
                target_df.iloc[target_row_idx, get_column_index('K', 'rid_library')] = "0x13"
                break

def update_target_without_source(subservice, target_df, target_row_idx, source_df, rid_start_row):
    """当 source 中没有匹配数据时，更新 target 数据
    Args:
        target_df: 目标数据框
        target_row_idx: 目标行索引
        subservice: 子服务类型 ('01', '02', '03')
        source_df: 源数据框
        rid_start_row: RID 在 source 中的起始行索引
    """
    # 更新 B 列 Description
    target_df.iloc[target_row_idx, get_column_index('B', 'rid_library')] = \
        source_df.iloc[rid_start_row, get_column_index('B', 'routine_control')]
    
    # 更新 C 列 APP 为 N
    target_df.iloc[target_row_idx, get_column_index('C', 'rid_library')] = "N"
    
    # 更新 D 列 Boot 为 N
    target_df.iloc[target_row_idx, get_column_index('D', 'rid_library')] = "N"
    
    # 更新 E 列 SubService
    target_df.iloc[target_row_idx, get_column_index('E', 'rid_library')] = f"0x{subservice}"
    
    # 其他列设置为 NA
    for col in range(get_column_index('F', 'rid_library'), len(target_df.columns)):
        target_df.iloc[target_row_idx, col] = "NA"

def process_subservice(source_df, target_df, rid, subservice):
    """处理单个 subservice 的数据"""
    # print(f"{'='*50}")
    # print(f"处理 RID {rid} 的 subservice {subservice}")
    # print(f"{'='*50}")

    # 1. 处理 target 数据
    target_row = target_df[
        target_df.iloc[:, get_column_index('A', 'routine_control')].astype(str).str.contains(rid, na=False)
    ]
    target_row_idx = target_row.index[0] + int(subservice) - 1
    
    # print(f"\nTarget 数据:")
    # print(target_df.iloc[target_row_idx])

    # 2. 处理 source 数据
    # 找到 RID 所在的行
    rid_start_row = source_df[source_df.iloc[:, get_column_index('A', 'routine_control')].astype(str) == rid].index[0]
    
    # 找到下一个 RID 的位置或文件末尾
    next_rid_row = len(source_df)
    for i in range(rid_start_row + 1, len(source_df)):
        if pd.notna(source_df.iloc[i, get_column_index('A', 'routine_control')]) and \
           str(source_df.iloc[i, get_column_index('A', 'routine_control')]).startswith('0x'):
            next_rid_row = i
            break

    # 在这个范围内查找匹配的子服务行
    matching_rows = []
    for row in range(rid_start_row, next_rid_row):
        g_value = str(source_df.iloc[row, get_column_index('G', 'routine_control')])
        if subservice in g_value:
            matching_rows.append(row)

    # # 3. 输出结果
    # if matching_rows:
    #     print(f"\nSource 匹配数据:")
    #     for idx, row in enumerate(matching_rows, 1):
    #         print(f"\n匹配项 {idx} (Row {row}):")
    #         print(source_df.iloc[row])
    # else:
    #     print(f"\n警告: 在源文件中未找到 RID {rid} 的 subservice {subservice} 的相关数据")

    # # 打印更新前的 target 数据
    # print(f"\n更新前的 target 数据:")
    # print(target_df.iloc[target_row_idx])

    # 4. 更新 target 数据
    if matching_rows:
        update_target_with_source_data(subservice,source_df, target_df, target_row_idx, matching_rows, rid_start_row)
    else:
        update_target_without_source(subservice, target_df, target_row_idx, source_df, rid_start_row)

    # # 打印更新后的 target 数据
    # print(f"\n更新后的 target 数据:")
    # print(target_df.iloc[target_row_idx])

    return

def main_process_31(source_df, target_df, rid):
    """处理单个 Routine Control ID 的数据"""
    
    # print(f"\nBegin to process {rid}")

    # 处理每个子服务
    for subservice in ['01', '02', '03']:
        process_subservice(
            source_df, 
            target_df, 
            rid, 
            subservice
        )
    
    
def process_31(source_file, target_file):
    print(f"Begin to process 31 sheet")
    source_df, target_df = pre_process_31(source_file, target_file)
    # 处理每个 Routine Control ID 的具体数据
    for idx in range(0, len(target_df), 3):  # 每三行处理一次
        rid = target_df.iloc[idx, 0]  # A列 Routine Control ID
        if pd.notna(rid):  # 确保不是空值
            main_process_31(source_df, target_df, rid)
    
    print(f"Finished process 31 sheet")
    return target_df